#!/usr/bin/env python3
import openpyxl
import json
import sys

# Excelファイルを開く
wb = openpyxl.load_workbook('評価用シート.xlsx')

# シート名を確認
print("Available sheets:", wb.sheetnames)
print("\n" + "="*80 + "\n")

# 正解データシートを読み込む
if '正解データ' in wb.sheetnames:
    ws = wb['正解データ']
    print("正解データシート:")
    print("-" * 80)

    # ヘッダー行を取得
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else "")
    print("Headers:", headers)
    print()

    # 全データを取得
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # 空行でない場合
            data.append(dict(zip(headers, row)))

    print(f"Total rows: {len(data)}")
    print("\nFirst 3 rows:")
    for i, row_data in enumerate(data[:3], 1):
        print(f"\nRow {i}:")
        for key, value in row_data.items():
            if value is not None and value != "":
                print(f"  {key}: {value}")

    # JSONとして出力
    with open('正解データ.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\nSaved to 正解データ.json")

    print("\n" + "="*80 + "\n")

# 期待イメージシートを読み込む
if '期待イメージ' in wb.sheetnames:
    ws = wb['期待イメージ']
    print("期待イメージシート:")
    print("-" * 80)

    # 全セルの内容を表示
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(row):  # 空行でない場合
            print(f"Row {i}: {[cell for cell in row if cell]}")
